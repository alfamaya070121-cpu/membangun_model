
import io
from io import BytesIO

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font

st.set_page_config(page_title="Dashboard Analisis Dual Cycle", layout="wide")

# ================================================================
# KONSTANTA DEFAULT (samakan dengan VBA)
# ================================================================
AMBANG_COMBO_MENIT_DEFAULT = 40
AMBANG_DUAL_MENIT_DEFAULT = 240  # = 4 jam (bukan 6 jam spt komentar VBA lama)
SIZE_ELIGIBLE_DEFAULT = 20


# ================================================================
# HELPER - baca & siapkan data
# ================================================================

def klasifikasi_activity(val: object) -> str:
    s = str(val).upper()
    return "LOAD" if "LOAD" in s else "DISC"


@st.cache_data(show_spinner=False)
def baca_file(file_bytes: bytes, filename: str, sheet_name=None):
    if filename.lower().endswith(".csv"):
        return {"__csv__": pd.read_csv(BytesIO(file_bytes))}
    xls = pd.ExcelFile(BytesIO(file_bytes))
    if sheet_name is not None:
        return {sheet_name: xls.parse(sheet_name)}
    return {name: xls.parse(name) for name in xls.sheet_names}


def siapkan_data(raw: pd.DataFrame, col_map: dict, size_eligible: int) -> pd.DataFrame:
    df = pd.DataFrame()
    df["VES_ID"] = raw[col_map["ves_id"]]
    df["CTR_SIZE"] = pd.to_numeric(raw[col_map["size"]], errors="coerce").fillna(0)
    df["CAR_CHE_ID"] = raw[col_map["truck"]].astype(str).str.strip()
    df["ACTIVITY"] = raw[col_map["activity"]].apply(klasifikasi_activity)
    df["TS_G"] = pd.to_datetime(raw[col_map["ts_g"]], errors="coerce")
    df["TS_H"] = pd.to_datetime(raw[col_map["ts_h"]], errors="coerce")

    # kalau G kosong pakai H, kalau H kosong pakai G (spt VBA)
    df["TS_G"] = df["TS_G"].fillna(df["TS_H"])
    df["TS_H"] = df["TS_H"].fillna(df["TS_G"])

    n_before = len(df)
    df = df.dropna(subset=["TS_G", "TS_H"]).reset_index(drop=True)
    n_dropped = n_before - len(df)
    if n_dropped:
        st.warning(
            f"{n_dropped} baris dibuang karena kedua kolom timestamp "
            f"(DISC_LOAD_TS & STACK_UNSTACK_TS) kosong / tidak valid."
        )

    df["ROW_IDX"] = df.index
    return df


# ================================================================
# LAYER 1 - COMBO / SINGLE
# ================================================================

def layer1_combo(df: pd.DataFrame, ambang_combo: float, size_eligible: int) -> pd.DataFrame:
    n = len(df)
    ts_g = df["TS_G"].to_numpy()
    ts_h = df["TS_H"].to_numpy()
    size = df["CTR_SIZE"].to_numpy()
    activity = df["ACTIVITY"].to_numpy()

    assigned = np.zeros(n, dtype=bool)
    group_id = np.zeros(n, dtype=int)

    pairs = []
    truck_positions = df.groupby("CAR_CHE_ID").indices  # dict truk -> array posisi (0..n-1)

    for _, pos in truck_positions.items():
        for act in ("LOAD", "DISC"):
            idx_act = [p for p in pos if activity[p] == act and size[p] == size_eligible]
            m = len(idx_act)
            for a in range(m - 1):
                i = idx_act[a]
                for b in range(a + 1, m):
                    k = idx_act[b]
                    gap_g = abs((ts_g[k] - ts_g[i]) / np.timedelta64(1, "m"))
                    gap_h = abs((ts_h[k] - ts_h[i]) / np.timedelta64(1, "m"))
                    gap = min(gap_g, gap_h)
                    if gap <= ambang_combo:
                        pairs.append((i, k, gap))

    # urutkan gap terkecil dulu, lalu greedy matching
    pairs.sort(key=lambda x: (x[2], x[0], x[1]))

    nxt = 0
    for i, k, _gap in pairs:
        if not assigned[i] and not assigned[k]:
            nxt += 1
            group_id[i] = nxt
            group_id[k] = nxt
            assigned[i] = assigned[k] = True

    for i in range(n):
        if not assigned[i]:
            nxt += 1
            group_id[i] = nxt
            assigned[i] = True

    out = df.copy()
    out["GROUP_ID"] = group_id
    return out


# ================================================================
# BENTUK EVENT
# ================================================================

def bentuk_event(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for gid, g in df.groupby("GROUP_ID"):
        activity = g["ACTIVITY"].iloc[0]
        truck = g["CAR_CHE_ID"].iloc[0]
        if activity == "DISC":
            starts, ends = g["TS_G"], g["TS_H"]
        else:
            starts, ends = g["TS_H"], g["TS_G"]
        container_status = "Combo" if len(g) >= 2 else "Single"
        records.append(
            {
                "GROUP_ID": gid,
                "ACTIVITY": activity,
                "CAR_CHE_ID": truck,
                "START_TS": starts.min(),
                "END_TS": ends.max(),
                "CONTAINER_STATUS": container_status,
            }
        )
    return pd.DataFrame(records)


# ================================================================
# LAYER 2 - DUAL CYCLE
# ================================================================

def layer2_dual(events: pd.DataFrame, ambang_dual: float) -> pd.DataFrame:
    events = events.reset_index(drop=True)
    n = len(events)
    start = events["START_TS"].to_numpy()
    end = events["END_TS"].to_numpy()
    activity = events["ACTIVITY"].to_numpy()

    assigned = np.zeros(n, dtype=bool)
    status = np.array(["Non Dual"] * n, dtype=object)

    pairs = []
    truck_positions = events.groupby("CAR_CHE_ID").indices

    for _, pos in truck_positions.items():
        pos_sorted = sorted(pos, key=lambda p: start[p])
        m = len(pos_sorted)
        for a in range(m - 1):
            i = pos_sorted[a]
            for b in range(a + 1, m):
                k = pos_sorted[b]
                gap_ab = (start[k] - end[i]) / np.timedelta64(1, "m")
                gap_ba = (start[i] - end[k]) / np.timedelta64(1, "m")
                if gap_ab >= 0:
                    gap = gap_ab
                elif gap_ba >= 0:
                    gap = gap_ba
                else:
                    gap = 0.0  # overlap waktu -> dianggap berdekatan

                if gap > ambang_dual:
                    # waktu mulai naik monoton -> aman berhenti di sini
                    break

                if activity[i] != activity[k]:
                    pairs.append((i, k, gap))

    pairs.sort(key=lambda x: (x[2], x[0], x[1]))
    for i, k, _gap in pairs:
        if not assigned[i] and not assigned[k]:
            status[i] = "Dual Cycle"
            status[k] = "Dual Cycle"
            assigned[i] = assigned[k] = True

    out = events.copy()
    out["STATUS"] = status
    return out


# ================================================================
# PENOMORAN EVENT_ID GLOBAL (urut truk sesuai kemunculan pertama
# di data asli, di dalam truk diurutkan berdasarkan START_TS)
# ================================================================

def beri_event_id(events: pd.DataFrame, df_asli: pd.DataFrame):
    truck_order = list(dict.fromkeys(df_asli["CAR_CHE_ID"].tolist()))
    rank = {tk: i for i, tk in enumerate(truck_order)}

    events = events.copy()
    events["_truck_rank"] = events["CAR_CHE_ID"].map(rank)
    events = events.sort_values(["_truck_rank", "START_TS"]).reset_index(drop=True)
    events["EVENT_ID"] = events.index + 1
    events = events.drop(columns=["_truck_rank"])

    event_id_map = dict(zip(events["GROUP_ID"], events["EVENT_ID"]))
    return events, event_id_map


def gabungkan_hasil(df: pd.DataFrame, events: pd.DataFrame, event_id_map: dict) -> pd.DataFrame:
    status_map = events.set_index("GROUP_ID")["STATUS"].to_dict()
    container_map = events.set_index("GROUP_ID")["CONTAINER_STATUS"].to_dict()

    out = df.copy()
    out["EVENT_ID"] = out["GROUP_ID"].map(event_id_map)
    out["CONTAINER_STATUS"] = out["GROUP_ID"].map(container_map)
    out["STATUS"] = out["GROUP_ID"].map(status_map)
    out = out.drop(columns=["GROUP_ID", "ROW_IDX"])
    return out


# ================================================================
# RINGKASAN / STATISTIK
# ================================================================

def hitung_ringkasan(events: pd.DataFrame, out_df: pd.DataFrame) -> dict:
    total_event = len(events)
    total_dual = int((events["STATUS"] == "Dual Cycle").sum())
    total_single = total_event - total_dual

    combo_dual = int(((events["CONTAINER_STATUS"] == "Combo") & (events["STATUS"] == "Dual Cycle")).sum())
    combo_single = int(((events["CONTAINER_STATUS"] == "Combo") & (events["STATUS"] == "Non Dual")).sum())
    single_dual = int(((events["CONTAINER_STATUS"] == "Single") & (events["STATUS"] == "Dual Cycle")).sum())
    single_single = int(((events["CONTAINER_STATUS"] == "Single") & (events["STATUS"] == "Non Dual")).sum())

    dual_load = int(((out_df["STATUS"] == "Dual Cycle") & (out_df["ACTIVITY"] == "LOAD")).sum())
    dual_disc = int(((out_df["STATUS"] == "Dual Cycle") & (out_df["ACTIVITY"] == "DISC")).sum())
    single_load = int(((out_df["STATUS"] == "Non Dual") & (out_df["ACTIVITY"] == "LOAD")).sum())
    single_disc = int(((out_df["STATUS"] == "Non Dual") & (out_df["ACTIVITY"] == "DISC")).sum())

    ev = events.copy()
    ev["BULAN"] = ev["START_TS"].dt.to_period("M")
    monthly = ev.groupby("BULAN").agg(
        total_event=("STATUS", "count"),
        dual=("STATUS", lambda s: int((s == "Dual Cycle").sum())),
    )
    monthly["non_dual"] = monthly["total_event"] - monthly["dual"]
    monthly["pct_dual"] = np.where(monthly["total_event"] > 0, monthly["dual"] / monthly["total_event"], 0)
    monthly = monthly.sort_index()
    monthly.index = monthly.index.astype(str)

    return {
        "total_event": total_event,
        "total_dual": total_dual,
        "total_single": total_single,
        "pct_dual": (total_dual / total_event) if total_event else 0,
        "combo_dual": combo_dual,
        "combo_single": combo_single,
        "single_dual": single_dual,
        "single_single": single_single,
        "dual_load": dual_load,
        "dual_disc": dual_disc,
        "single_load": single_load,
        "single_disc": single_disc,
        "n_truck": out_df["CAR_CHE_ID"].nunique(),
        "monthly": monthly,
    }


# ================================================================
# EXPORT EXCEL (Data + Ringkasan + Chart), mirip output VBA
# ================================================================

def build_excel_download(out_df: pd.DataFrame, summary: dict, ambang_combo: float, ambang_dual: float) -> BytesIO:
    wb = Workbook()

    # ---- Sheet Data ----
    ws_data = wb.active
    ws_data.title = "Data"

    export_df = out_df.copy()
    for c in export_df.columns:
        if pd.api.types.is_datetime64_any_dtype(export_df[c]):
            export_df[c] = export_df[c].dt.strftime("%Y-%m-%d %H:%M:%S")

    ws_data.append(list(export_df.columns))
    for row in export_df.itertuples(index=False):
        ws_data.append(list(row))
    for cell in ws_data[1]:
        cell.font = Font(bold=True)
    for col_cells in ws_data.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_data.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    # ---- Sheet Ringkasan ----
    ws = wb.create_sheet("Ringkasan Dual Cycle")
    ws["A1"] = "RINGKASAN ANALISIS DUAL CYCLE"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Ambang Combo: {ambang_combo:g} menit | Ambang Dual Cycle: {ambang_dual:g} menit | "
        f"Basis perhitungan: EVENT_ID (ritase) | Pairing dibatasi per truk (CAR_CHE_ID)"
    )
    ws["A2"].font = Font(italic=True)

    ws["A4"] = "RINGKASAN (BERBASIS EVENT_ID / RITASE)"
    ws["A4"].font = Font(bold=True)
    ws["A5"], ws["B5"] = "Total Event (Ritase)", summary["total_event"]
    ws["A6"], ws["B6"] = "Dual Cycle", summary["total_dual"]
    ws["A6"].font = Font(bold=True)
    ws["B6"].font = Font(bold=True)
    ws["A7"], ws["B7"] = "Non Dual", summary["total_single"]
    ws["A8"], ws["B8"] = "Persentase Dual Cycle", summary["pct_dual"]
    ws["B8"].number_format = "0.0%"
    ws["A8"].font = Font(bold=True)
    ws["B8"].font = Font(bold=True)

    ws["A10"] = "RINCIAN CONTAINER (BERBASIS EVENT_ID)"
    ws["A10"].font = Font(bold=True)
    ws["A11"], ws["B11"] = "Combo - Dual Cycle", summary["combo_dual"]
    ws["A12"], ws["B12"] = "Combo - Non Dual", summary["combo_single"]
    ws["A13"], ws["B13"] = "Single - Dual Cycle", summary["single_dual"]
    ws["A14"], ws["B14"] = "Single - Non Dual", summary["single_single"]

    ws["A16"] = "RINCIAN AKTIVITAS (BERBASIS BARIS LOAD/DISC -- INFO TAMBAHAN)"
    ws["A16"].font = Font(bold=True)
    ws["A17"], ws["B17"] = "Total aktivitas (baris)", len(out_df)
    ws["A18"], ws["B18"] = "Dual Cycle - LOAD", summary["dual_load"]
    ws["A19"], ws["B19"] = "Dual Cycle - DISC", summary["dual_disc"]
    ws["A20"], ws["B20"] = "Non Dual - LOAD", summary["single_load"]
    ws["A21"], ws["B21"] = "Non Dual - DISC", summary["single_disc"]
    ws["A23"], ws["B23"] = "Jumlah truk (CAR_CHE_ID) terdeteksi", summary["n_truck"]

    ws["A25"] = "BREAKDOWN BULANAN (EVENT / RITASE)"
    ws["A25"].font = Font(bold=True)
    header_row = 26
    for c, label in zip("ABCDE", ["Bulan", "Total Event", "Dual Cycle", "Non Dual", "% Dual Cycle"]):
        ws[f"{c}{header_row}"] = label
        ws[f"{c}{header_row}"].font = Font(bold=True)

    monthly = summary["monthly"]
    r = header_row
    for bulan, row in monthly.iterrows():
        r += 1
        ws.cell(r, 1, str(bulan))
        ws.cell(r, 2, int(row["total_event"]))
        ws.cell(r, 3, int(row["dual"]))
        ws.cell(r, 4, int(row["non_dual"]))
        ws.cell(r, 5, float(row["pct_dual"]))
        ws.cell(r, 5).number_format = "0.0%"

    total_row = r + 1
    tot_evt = int(monthly["total_event"].sum()) if len(monthly) else 0
    tot_dual = int(monthly["dual"].sum()) if len(monthly) else 0
    ws.cell(total_row, 1, "Total")
    ws.cell(total_row, 2, tot_evt)
    ws.cell(total_row, 3, tot_dual)
    ws.cell(total_row, 4, tot_evt - tot_dual)
    ws.cell(total_row, 5, (tot_dual / tot_evt) if tot_evt else 0)
    ws.cell(total_row, 5).number_format = "0.0%"
    for c in range(1, 6):
        ws.cell(total_row, c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 46
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    # ---- Pie chart: Dual vs Non Dual ----
    pie = PieChart()
    pie.title = "Dual Cycle vs Non Dual (Event)"
    data = Reference(ws, min_col=2, min_row=6, max_row=7)
    cats = Reference(ws, min_col=1, min_row=6, max_row=7)
    pie.add_data(data)
    pie.set_categories(cats)
    ws.add_chart(pie, "G4")

    # ---- Bar chart: Container x Status ----
    bar = BarChart()
    bar.type = "col"
    bar.title = "Container x Status (Event)"
    data = Reference(ws, min_col=2, min_row=11, max_row=14)
    cats = Reference(ws, min_col=1, min_row=11, max_row=14)
    bar.add_data(data)
    bar.set_categories(cats)
    ws.add_chart(bar, "G20")

    # ---- Stacked bar: breakdown bulanan ----
    if len(monthly) > 0:
        bar2 = BarChart()
        bar2.type = "col"
        bar2.grouping = "stacked"
        bar2.overlap = 100
        bar2.title = "Breakdown Bulanan: Dual Cycle vs Non Dual (Event)"
        data = Reference(ws, min_col=3, max_col=4, min_row=header_row, max_row=r)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=r)
        bar2.add_data(data, titles_from_data=True)
        bar2.set_categories(cats)
        ws.add_chart(bar2, "G36")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ================================================================
# ================================================================
# UI STREAMLIT
# ================================================================
# ================================================================

st.title("🚛 Dashboard Analisis Dual Cycle")
st.caption(
    "Port dari macro VBA Analisis Dual Cycle. Upload data mentah, atur ambang batas, "
    "lihat hasilnya dalam chart interaktif, lalu download hasil analisis lengkap."
)

with st.sidebar:
    st.header("⚙️ Pengaturan")

    ambang_combo = st.number_input(
        "Ambang Combo (menit)", min_value=1, value=AMBANG_COMBO_MENIT_DEFAULT, step=5,
        help="Jarak waktu maksimum antar 2 baris size 20ft, truk & aktivitas sama, supaya dianggap 'Combo'.",
    )
    ambang_dual = st.number_input(
        "Ambang Dual Cycle (menit)", min_value=1, value=AMBANG_DUAL_MENIT_DEFAULT, step=10,
        help="Jarak waktu maksimum antar 2 event beda aktivitas (LOAD vs DISC) dalam truk yang sama.",
    )
    size_eligible = st.number_input(
        "Ukuran kontainer eligible untuk Combo (ft)", min_value=1, value=SIZE_ELIGIBLE_DEFAULT, step=5,
    )

st.subheader("1️⃣ Upload Data")
uploaded = st.file_uploader("Upload file .xlsx atau .csv", type=["xlsx", "xls", "csv"])

if uploaded is None:
    st.info("Silakan upload file data untuk memulai analisis.")
    st.stop()

file_bytes = uploaded.getvalue()
sheets = baca_file(file_bytes, uploaded.name)

sheet_name = None
if len(sheets) > 1:
    sheet_name = st.selectbox("Pilih sheet data", list(sheets.keys()))
else:
    sheet_name = list(sheets.keys())[0]

raw = sheets[sheet_name]
st.write(f"Preview data mentah ({len(raw)} baris):")
st.dataframe(raw.head(10), use_container_width=True)

st.subheader("2️⃣ Pemetaan Kolom")
cols = list(raw.columns)


def guess(options, keywords, default_idx=0):
    for kw in keywords:
        for i, c in enumerate(options):
            if kw.lower() in str(c).lower():
                return i
    return default_idx


c1, c2, c3 = st.columns(3)
with c1:
    col_ves = st.selectbox("VES_ID", cols, index=guess(cols, ["ves"])))
    col_size = st.selectbox("CTR_SIZE (ukuran kontainer)", cols, index=guess(cols, ["size", "ctr_size"]))
with c2:
    col_truck = st.selectbox("CAR_CHE_ID (truk)", cols, index=guess(cols, ["car_che", "truck", "che"]))
    col_activity = st.selectbox("ACTIVITY", cols, index=guess(cols, ["activity", "aktivitas"]))
with c3:
    col_ts_g = st.selectbox("DISC_LOAD_TS", cols, index=guess(cols, ["disc_load", "disc_loading"]))
    col_ts_h = st.selectbox("STACK_UNSTACK_TS", cols, index=guess(cols, ["stack_unstack", "unstack_stack"]))

col_map = {
    "ves_id": col_ves,
    "size": col_size,
    "truck": col_truck,
    "activity": col_activity,
    "ts_g": col_ts_g,
    "ts_h": col_ts_h,
}

run = st.button("▶️ Jalankan Analisis Dual Cycle", type="primary")

if not run and "hasil" not in st.session_state:
    st.stop()

if run:
    with st.spinner("Memproses data..."):
        df = siapkan_data(raw, col_map, size_eligible)
        df_combo = layer1_combo(df, ambang_combo, size_eligible)
        events = bentuk_event(df_combo)
        events = layer2_dual(events, ambang_dual)
        events, event_id_map = beri_event_id(events, df_combo)
        out_df = gabungkan_hasil(df_combo, events, event_id_map)
        summary = hitung_ringkasan(events, out_df)

    st.session_state["hasil"] = {
        "out_df": out_df,
        "events": events,
        "summary": summary,
        "ambang_combo": ambang_combo,
        "ambang_dual": ambang_dual,
    }

hasil = st.session_state["hasil"]
out_df = hasil["out_df"]
events = hasil["events"]
summary = hasil["summary"]

st.success("Analisis selesai!")

# ================================================================
# RINGKASAN / KPI
# ================================================================

st.subheader("3️⃣ Ringkasan Hasil")

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Total Event (Ritase)", summary["total_event"])
k2.metric("Dual Cycle", summary["total_dual"])
k3.metric("Non Dual", summary["total_single"])
k4.metric("% Dual Cycle", f"{summary['pct_dual']*100:.1f}%")
k5.metric("Jumlah Truk", summary["n_truck"])

# ================================================================
# CHART
# ================================================================

cc1, cc2 = st.columns(2)

with cc1:
    pie_df = pd.DataFrame(
        {"Status": ["Dual Cycle", "Non Dual"], "Jumlah": [summary["total_dual"], summary["total_single"]]}
    )
    fig_pie = px.pie(
        pie_df, names="Status", values="Jumlah", hole=0.45,
        title="Dual Cycle vs Non Dual (berbasis Event)",
        color="Status",
        color_discrete_map={"Dual Cycle": "#2E86AB", "Non Dual": "#E76F51"},
    )
    fig_pie.update_traces(textinfo="percent+label")
    st.plotly_chart(fig_pie, use_container_width=True)

with cc2:
    container_df = pd.DataFrame(
        {
            "Container": ["Combo", "Combo", "Single", "Single"],
            "Status": ["Dual Cycle", "Non Dual", "Dual Cycle", "Non Dual"],
            "Jumlah": [
                summary["combo_dual"], summary["combo_single"],
                summary["single_dual"], summary["single_single"],
            ],
        }
    )
    fig_bar = px.bar(
        container_df, x="Container", y="Jumlah", color="Status", barmode="group",
        title="Rincian Container x Status (Event)",
        color_discrete_map={"Dual Cycle": "#2E86AB", "Non Dual": "#E76F51"},
        text="Jumlah",
    )
    st.plotly_chart(fig_bar, use_container_width=True)

monthly = summary["monthly"].reset_index().rename(columns={"BULAN": "Bulan"})
if len(monthly) > 0:
    monthly_long = monthly.melt(
        id_vars="Bulan", value_vars=["dual", "non_dual"], var_name="Status", value_name="Jumlah"
    )
    monthly_long["Status"] = monthly_long["Status"].map({"dual": "Dual Cycle", "non_dual": "Non Dual"})
    fig_month = px.bar(
        monthly_long, x="Bulan", y="Jumlah", color="Status", barmode="stack",
        title="Breakdown Bulanan: Dual Cycle vs Non Dual (Event)",
        color_discrete_map={"Dual Cycle": "#2E86AB", "Non Dual": "#E76F51"},
    )
    fig_line = go.Scatter(
        x=monthly["Bulan"], y=monthly["pct_dual"] * 100, name="% Dual Cycle",
        yaxis="y2", mode="lines+markers", line=dict(color="#F4A261", width=3),
    )
    fig_month.add_trace(fig_line)
    fig_month.update_layout(
        yaxis2=dict(title="% Dual Cycle", overlaying="y", side="right", range=[0, 100]),
    )
    st.plotly_chart(fig_month, use_container_width=True)

with st.expander("📋 Rincian Aktivitas (berbasis baris LOAD/DISC — info tambahan)"):
    act_df = pd.DataFrame(
        {
            "Status": ["Dual Cycle", "Dual Cycle", "Non Dual", "Non Dual"],
            "Activity": ["LOAD", "DISC", "LOAD", "DISC"],
            "Jumlah": [
                summary["dual_load"], summary["dual_disc"],
                summary["single_load"], summary["single_disc"],
            ],
        }
    )
    st.dataframe(act_df, use_container_width=True, hide_index=True)

# ================================================================
# DATA HASIL
# ================================================================

st.subheader("4️⃣ Data Hasil Analisis")
st.dataframe(out_df, use_container_width=True, height=400)

# ================================================================
# DOWNLOAD
# ================================================================

st.subheader("5️⃣ Download Hasil")

excel_bio = build_excel_download(out_df, summary, hasil["ambang_combo"], hasil["ambang_dual"])

dcol1, dcol2 = st.columns(2)
with dcol1:
    st.download_button(
        "⬇️ Download Excel (Data + Ringkasan + Chart)",
        data=excel_bio,
        file_name="Hasil_Analisis_Dual_Cycle.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with dcol2:
    csv_bytes = out_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "⬇️ Download CSV (Data saja)",
        data=csv_bytes,
        file_name="Hasil_Analisis_Dual_Cycle.csv",
        mime="text/csv",
        use_container_width=True,
    )
